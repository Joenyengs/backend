from datetime import date
from statistics import mean
from rest_framework.generics import RetrieveUpdateAPIView, CreateAPIView, ListAPIView, RetrieveAPIView
from rest_framework.permissions import IsAuthenticated, IsAdminUser, AllowAny
from rest_framework.views import APIView
from rest_framework import status
from rest_framework.response import Response
from rest_framework import viewsets, permissions
from users.models import Notification
from django.http import JsonResponse, HttpResponse
from .models import Candidature, QuizAnswer, Recours, RecoursActionHistory, TrainingModule, QuizQuestion, Traitement
from .serializers import CandidatureAddSerializer, CandidatureFrontendSerializer, CandidatureStatusSerializer, ListeCandidatureSerializer, QuizAnswerSerializer, RecoursActionHistorySerializer, RecoursSerializer, RecoursStatusSerializer, TrainingModuleSerializer, QuizQuestionSerializer, TraitementListSerializer, TraitementSerializer, TraitementWithCandidatureSerializer, TraiterCandidatureSerializer
from django.utils.timezone import now
from django.core.exceptions import ValidationError as DjangoValidationError
from rest_framework.exceptions import ValidationError as DRFValidationError
from rest_framework.decorators import action
from .permissions import IsAdmin, IsAlumni, IsEvaluator, IsTeacher, IsStudent, IsCandidat
from django.core.mail import send_mail
from rest_framework_simplejwt.authentication import JWTAuthentication
from django.db.models import Count, Q, F, ExpressionWrapper, IntegerField
import openpyxl
from openpyxl.utils import get_column_letter
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from users.models import profile_completion_ratio
from django.contrib.auth import get_user_model
from openpyxl.styles import Font


################################### CANDIDATURES ###############################
class CandidatureCreateView(CreateAPIView):
    queryset = Candidature.objects.all()
    serializer_class = CandidatureAddSerializer
    permission_classes = [IsCandidat]
    authentication_classes = [JWTAuthentication]

    def create(self, request, *args, **kwargs):
        user = request.user
        if not user.role == 'candidat':
            return Response({'Erreur': 'Cette opération requiert uniquement les privilèges Candidat.'}, status=status.HTTP_403_FORBIDDEN)

        if Candidature.objects.filter(candidat=user).exists():
            Response({'Erreur': 'Vous avez déjà soumis une candidature.'}, status=status.HTTP_403_FORBIDDEN)

        user_profil = user.profil_candidat

        # rejete la requete si le profil n'est pas renseigné à au moins 80%
        # if not profile_completion_ratio(user_profil)['is_80_percent_or_more']:
        #     return Response({'Erreur': 'Votre profil doit être complété à au moins 80% pour soumettre une candidature.'}, status=status.HTTP_400_BAD_REQUEST)

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            self.perform_create(serializer)
        except Exception as e:
            return Response({'Erreur': f'Erreur lors de la création de la candidature: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        # Envoi d'une notification à l'utilisateur
        n = Notification.objects.create(
            message="Votre candidature a été créée avec succès.",
            link="/espace-candidat"
        )
        n.users.set([user])

        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def perform_create(self, serializer):
        serializer.save(candidat=self.request.user)


class ListeCandidaturesView(ListAPIView):
    serializer_class = CandidatureFrontendSerializer # ListeCandidatureSerializer
    permission_classes = [IsAdminUser | IsEvaluator]
    authentication_classes = [JWTAuthentication]

    def get_queryset(self):
        user = self.request.user

        if user.is_staff or user.is_superuser or user.role in ['admin', 'evaluateur']:
            # Paginate the queryset in subsets of 50 objects each
            page_size = 50
            page = int(self.request.query_params.get('page', 1))
            start = (page - 1) * page_size
            end = start + page_size
            return Candidature.objects.all()[start:end]
        
        # Sinon, on retourne uniquement la candidature liée à l'utilisateur
        return Candidature.objects.filter(candidat=user)
        
        # Si l'utilisateur est un évaluateur, on retourne les candidatures qu'il a traitées
        # if user.role in ['evaluateur']:
        #     result = self.recharger(self.request)
        #     if result is not None:
                # result is a Response, but get_queryset must return a queryset
                # You probably want to return the candidatures assigned to this evaluator
                # return result   #Candidature.objects.filter(traitements__evaluateur=user).distinct()
            # Use the result if necessary, e.g., logging or further processing
            # print(result)  # Example usage
            # return Candidature.objects.filter(traite_par=user)
        # return Candidature.objects.none()

    # def get_queryset(self):
    #     user = self.request.user

    #     if user.is_staff or user.is_superuser or user.role == 'admin':
    #         # Paginate the queryset in subsets of 50 objects each
    #         page_size = 50
    #         page = int(self.request.query_params.get('page', 1))
    #         start = (page - 1) * page_size
    #         end = start + page_size
    #         return Candidature.objects.all()[start:end]
        
    #     # Si l'utilisateur est un candidat, on retourne uniquement sa candidature
    #     if user.role == 'candidat':
    #         return Candidature.objects.filter(candidat=user)
        
    #     # Si l'utilisateur est un évaluateur, on retourne les candidatures qu'il a traitées
    #     if user.role in ['evaluateur']:
    #         result = self.recharger(self.request)
    #         if result is not None:
    #             # result is a Response, but get_queryset must return a queryset
    #             # You probably want to return the candidatures assigned to this evaluator
    #             return result   #Candidature.objects.filter(traitements__evaluateur=user).distinct()
    #         # Use the result if necessary, e.g., logging or further processing
    #         # print(result)  # Example usage
    #         # return Candidature.objects.filter(traite_par=user)
    #     return Candidature.objects.none()
        # No additional methods needed here for your logic.
    # @action(detail=False, methods=['get'], permission_classes=[IsEvaluator])
    # def recharger(self, request):
    #     """
    #     Permet à un évaluateur de recharger sa file d'attente avec un nouveau lot de candidatures.
    #     """
    #     user = request.user
    #     lot_size = int(request.data.get('lot_size', 50))  # Taille du lot à attribuer

    #     # Récupère les candidatures disponibles (pas plus de 2 traitements, pas encore traitées par cet évaluateur)
    #     candidatures_disponibles = (
    #         Candidature.objects
    #         .annotate(nb_traitements=Count('traitements'))
    #         .filter(
    #             nb_traitements__lt=2,
    #             # traitements__evaluateur__ne=user  # Exclure celles déjà traitées par cet évaluateur
    #         )
    #         .exclude(traitements__evaluateur=user)
    #         .distinct()
    #     )

    #     # Sélectionne un nouveau lot
    #     lot = candidatures_disponibles[:lot_size]

    #     # Pour chaque candidature du lot, crée une entrée Traitement si non existante pour cet évaluateur
    #     for candidature in lot:
    #         if not Traitement.objects.filter(candidature=candidature, evaluateur=user).exists():
    #             Traitement.objects.create(
    #                 candidature=candidature,
    #                 evaluateur=user,
    #                 # Ajoutez d'autres champs par défaut si nécessaire
    #             )

    #     return [CandidatureFrontendSerializer(c).data for c in lot]
       

class CandidatureViewSet(viewsets.ModelViewSet):
    """
    Permet :
        (1) au candidat de gérer sa candidature : création, modification, envoi.
        (2) à l'administrateur de consulter les candidatures.
    """
    queryset = Candidature.objects.all()
    serializer_class = CandidatureAddSerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.is_staff or user.is_superuser:
            return Candidature.objects.all()
        if user.role == 'candidat':
            # Si l'utilisateur est un candidat, on retourne uniquement sa candidature
            return Candidature.objects.filter(candidat=user)
        if user.role in ['evaluateur']:
            self.recharger(self.request)
            # Si l'utilisateur est un évaluateur, on retourne les candidatures qu'il a traitées
            # return Candidature.objects.filter(traite_par=user)
        return None

    def perform_create(self, serializer):

        try:
            serializer.save(candidat=self.request.user)
        except Exception as e:
            return Response({'Erreur': f'Erreur lors de la création de la candidature: {str(e)}' }, status=status.HTTP_400_BAD_REQUEST)
            # raise DRFValidationError({'Erreur': f'Erreur lors de la création de la candidature: {str(e)}'})

    def perform_update(self, serializer):
        instance = serializer.instance
        if instance.statut != 'non_envoye':
            return Response({'restriction': 'Cette candidature ne peut plus être modifiée.'}, status=403)
        serializer.save()

    # pour les évaluateurs, on peut ajouter une action pour recharger leur file d'attente
    @action(detail=False, methods=['get'], permission_classes=[IsEvaluator])
    def recharger(self, request):
        """
        Permet à un évaluateur de recharger sa file d'attente avec un nouveau lot de candidatures.
        """
        user = request.user
        lot_size = int(request.query_params.get('lot_size', 50))  # Taille du lot à attribuer

        # Récupère les candidatures disponibles (pas plus de 2 traitements, pas déjà traitées par cet évaluateur)
        candidatures_disponibles = (
            Candidature.objects
            .annotate(nb_traitements=Count('traitements'))
            .filter(
                nb_traitements__lt=3,  # Au moins 2 traitements requis
                #traitements__evaluateur__ne=user  # Exclure celles déjà traitées par cet évaluateur
            )
            .exclude(traitements__evaluateur=user)
            .distinct()
        )

        # Sélectionne un nouveau lot
        lot = candidatures_disponibles[:lot_size]

        return Response([CandidatureFrontendSerializer(c).data for c in lot])


class CandidatureStatusView(RetrieveAPIView):
    """
    Vue pour récupérer le statut de la candidature d'un utilisateur.
    Cette vue permet à :
        (1) un utilisateur authentifié d'obtenir le statut actuel de sa candidature
        (2) à l'administreur de lister les candidatures et les regrouper par statut
    Utilise l'authentification JWT et nécessite que l'utilisateur soit authentifié.
    Méthodes :
        get_object() : Retourne l'instance de Candidature associée à l'utilisateur courant.
    Sérialiseur :
        CandidatureStatusSerializer
    Authentification :
        JWTAuthentication
    Permissions :
        IsAuthenticated
    """

    serializer_class = CandidatureStatusSerializer
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get_object(self):
        user = self.request.user
        if user.is_superuser:
            return Candidature.objects.all()
        return Candidature.objects.get(candidat=user)
    
    def get(self, request, *args, **kwargs):
        """
        Retourne le statut de la candidature de l'utilisateur courant.
        Si l'utilisateur n'a pas de candidature, retourne un message d'erreur.
        """
        try:
            candidature = self.get_object()
            serializer = self.get_serializer(candidature, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Candidature.DoesNotExist:
            return Response({'detail': 'Aucune candidature trouvée pour cet utilisateur.'}, status=status.HTTP_404_NOT_FOUND)


class StatsCandidaturesParProvinceParGenreView(APIView):
    permission_classes = [IsAdminUser]
    authentication_classes = [JWTAuthentication]
    queryset = Candidature.objects.all()

    def get(self, request):
        # Statistiques par province avec total hommes, femmes et total général
        stats = (
            Candidature.objects
            .values(province=F('candidat__profil_candidat__province_de_residence'))
            .annotate(
            total_hommes=Count('id', filter=Q(candidat__profil_candidat__genre='M')),
            total_femmes=Count('id', filter=Q(candidat__profil_candidat__genre='F')),
            total=Count('id')
            )
            .order_by('province')
        )

        return Response(stats, status=status.HTTP_200_OK)
    

class ExportCandidatureStatsExcelView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        stats = (
            Candidature.objects
            .values('candidat__province_de_residence')
            .annotate(
                total_hommes=Count('id', filter=Q(candidat__genre='M')),
                total_femmes=Count('id', filter=Q(candidat__genre='F')),
                total=Count('id')
            )
            .order_by('candidat__province_de_residence')
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stats Candidatures"

        ws.append(["Province", "Hommes", "Femmes", "Total"])
        for row in stats:
            ws.append([
                row['candidat__province_de_residence'],
                row['total_hommes'],
                row['total_femmes'],
                row['total']
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=stats_candidatures_par_province.xlsx'
        wb.save(response)
        return response


class StatsCandidaturesTraiteesParPremierEvaluateurView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        # Récupère les candidatures qui ont été traitées une seule fois (un seul Traitement lié)
        stats = (
            Candidature.objects
            .annotate(nb_traitements=Count('traitements'))
            .filter(nb_traitements=1)
            .values('id', 'candidat__first_name', 'candidat__last_name', 'candidat__email', 'traitements__evaluateur__id', 'traitements__evaluateur__email')
            .order_by('id')
        )
        return Response({
            "traitees_une_fois": stats
        }, status=status.HTTP_200_OK)
    

class ExportCandidatureTraiteesParPremierEvaluateurExcelView(APIView):
    permission_classes = [IsAdminUser]
    # authentication_classes = []

    def get(self, request):
        stats = (
            Candidature.objects
            .annotate(nb_traitements=Count('traitements'))
            .filter(nb_traitements=1)
            .values('id', 'candidat__first_name', 'candidat__last_name', 'candidat__email', 'traitements__evaluateur__id', 'traitements__evaluateur__email')
            .order_by('id')
        )


        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stats Candidatures Traitées par Premier Evaluateur"

        # En-tête en gras
        bold_font = Font(bold=True)
        header = ["Prénom Candidat", "Nom Candidat", "Email Candidat", "Email Evaluateur"]
        ws.append(header)
        for col_num, _ in enumerate(header, 1):
            ws.cell(row=1, column=col_num).font = bold_font

        for row in stats:
            ws.append([
                # row['traite_par__id'],
                row['candidat__first_name'],
                row['candidat__last_name'],
                row['candidat__email'],
                row['traitements__evaluateur__email']
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=stats_candidatures_traitees_premier_evaluateur.xlsx'
        wb.save(response)
        return (response)


class StatsCandidaturesTraiteesParSecondEvaluateurView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
         # Récupère les candidatures qui ont été traitées deux fois (deux entrées Traitement)
        candidatures = (
            Candidature.objects
            .annotate(nb_traitements=Count('traitements'))
            .filter(nb_traitements=2)
            .prefetch_related('traitements__evaluateur')
        )

        stats = []
        for c in candidatures:
            evaluateurs = []
            for t in c.traitements.all():
                evaluateurs.append(t.evaluateur.email)
            stats.append({
                'id': c.id,
                'candidat.first_name': c.candidat.first_name,
                'candidat.last_name': c.candidat.last_name,
                'candidat.email': c.candidat.email,
                'evaluateurs': evaluateurs,
            })
        
        return Response({
            "traitees_deux_fois": stats
        }, status=status.HTTP_200_OK)


class ExportCandidatureTraiteesParSecondEvaluateurExcelView(APIView):
    permission_classes = [IsAdminUser]
    # authentication_classes = []

    def get(self, request):
        candidatures = (
            Candidature.objects
            .annotate(nb_traitements=Count('traitements'))
            .filter(nb_traitements=2)
            .prefetch_related('traitements__evaluateur')
        )

        stats = []
        for c in candidatures:
            evaluateurs = []
            for t in c.traitements.all():
                evaluateurs.append(t.evaluateur.email)
            stats.append({
                'id': c.id,
                'candidat__first_name': c.candidat.first_name,
                'candidat__last_name': c.candidat.last_name,
                'candidat__email': c.candidat.email,
                'evaluateurs': evaluateurs,
            })

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stats Candidatures Traitées par Second Evaluateur"

        # En-tête en gras
        bold_font = Font(bold=True)
        header = ["Prénom Candidat", "Nom Candidat", "Email Candidat", "Email Evaluateur"]
        ws.append(header)
        for col_num, _ in enumerate(header, 1):
            ws.cell(row=1, column=col_num).font = bold_font

        for row in stats:
            # Insert all evaluateur emails in a single cell, separated by commas
            ws.append([
                row['candidat__first_name'],
                row['candidat__last_name'],
                row['candidat__email'],
                ", ".join(row['evaluateurs'])
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=stats_candidatures_traitees_second_evaluateur.xlsx'
        wb.save(response)
        return (response)


class StatsCandidaturesTraiteesParTroisiemeEvaluateurView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
         # Récupère les candidatures qui ont été traitées trois fois (trois entrées Traitement)
        candidatures = (
            Candidature.objects
            .annotate(nb_traitements=Count('traitements'))
            .filter(nb_traitements=3)
            .prefetch_related('traitements__evaluateur')
        )

        stats = []
        for c in candidatures:
            evaluateurs = []
            for t in c.traitements.all():
                evaluateurs.append(t.evaluateur.email)
            stats.append({
                'id': c.id,
                'candidat.first_name': c.candidat.first_name,
                'candidat.last_name': c.candidat.last_name,
                'candidat.email': c.candidat.email,
                'evaluateurs': evaluateurs,
            })
        
        return Response({
            "traitees_trois_fois": stats
        }, status=status.HTTP_200_OK)


class ExportCandidatureTraiteesParTroisiemeEvaluateurExcelView(APIView):
    permission_classes = [IsAdminUser]
    # authentication_classes = []

    def get(self, request):
        candidatures = (
            Candidature.objects
            .annotate(nb_traitements=Count('traitements'))
            .filter(nb_traitements=3)
            .prefetch_related('traitements__evaluateur')
        )

        stats = []
        for c in candidatures:
            evaluateurs = []
            for t in c.traitements.all():
                evaluateurs.append(t.evaluateur.email)
            stats.append({
                'id': c.id,
                'candidat__first_name': c.candidat.first_name,
                'candidat__last_name': c.candidat.last_name,
                'candidat__email': c.candidat.email,
                'evaluateurs': evaluateurs,
            })

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stats Candidatures Traitées par Troisieme Evaluateur"

        # En-tête en gras
        bold_font = Font(bold=True)
        header = ["Prénom Candidat", "Nom Candidat", "Email Candidat", "Email Evaluateur"]
        ws.append(header)
        for col_num, _ in enumerate(header, 1):
            ws.cell(row=1, column=col_num).font = bold_font

        for row in stats:
            # Insert all evaluateur emails in a single cell, separated by commas
            ws.append([
                row['candidat__first_name'],
                row['candidat__last_name'],
                row['candidat__email'],
                ", ".join(row['evaluateurs'])
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=stats_candidatures_traitees_troisieme_evaluateur.xlsx'
        wb.save(response)
        return (response)



class StatsCandidaturesNonTraiteesView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        # Filtre les candidatures non traitées
        non_traitees = Candidature.objects.filter(statut__in=['envoye'])

        # Statistiques des candidatures exclues pour la nationalité
        stats_exclusions_nationalite = (
            non_traitees
            .exclude(candidat__profil_candidat__nationalite="RDC")
            .values('candidat__profil_candidat__nationalite')
            .annotate(total=Count('id'))
            .order_by('-total')
        )

        # Statistiques des candidatures exclues pour niveau d'étude (inférieur à bac+5)
        stats_exclusions_bac5 = (
            non_traitees
            .filter(
                Q(candidat__profil_candidat__niveau_etude__iexact='licence_bac+3') |
                Q(candidat__profil_candidat__niveau_etude__iexact='graduat') |
                Q(candidat__profil_candidat__niveau_etude__iexact='diplome_etat')
            )
            .count()
        )

        # Statistiques par tranche d'âge (exemple : <25, 25-30, 31-35, >35)
        today = date.today()
        age_expr = ExpressionWrapper(
            today.year - F('candidat__profil_candidat__date_de_naissance__year'),
            output_field=IntegerField()
        )
        stats_exclusions_age = (
            non_traitees
            .annotate(age=age_expr)
            .filter(Q(age__lt=18) | Q(age__gt=35))
            .count()
        )
        ages = non_traitees.annotate(age=age_expr).values_list('age', flat=True)

        tranches = {'<25': 0, '25-30': 0, '31-35': 0, '>35': 0}
        for age in ages:
            if age is None:
                continue
            if age < 25:
                tranches['<25'] += 1
            elif 25 <= age <= 30:
                tranches['25-30'] += 1
            elif 31 <= age <= 35:
                tranches['31-35'] += 1
            else:
                tranches['>35'] += 1

        return Response({
            "exclusions_nationalite": stats_exclusions_nationalite,
            "exclusions_niveau_etude_inferieur_bac+5": stats_exclusions_bac5,
            "exclusions_age": stats_exclusions_age,
            "tranches_age": tranches,
        })


class CandidaturesRetenuesStatsView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        # Filtre les candidatures retenues
        retenues = Candidature.objects.filter(statut='valide')

        # Statistiques par province
        stats_province = (
            retenues
            .values('candidat__profil_candidat__province_residence')
            .annotate(total=Count('id'))
            .order_by('candidat__profil_candidat__province_residence')
        )

        # Statistiques par genre
        stats_genre = (
            retenues
            .values('candidat__profil_candidat__genre')
            .annotate(total=Count('id'))
            .order_by('candidat__profil_candidat__genre')
        )

        # Statistiques par niveau d'étude
        stats_niveau = (
            retenues
            .values('candidat__profil_candidat__niveau_etude')
            .annotate(total=Count('id'))
            .order_by('candidat__profil_candidat__niveau_etude')
        )

        return Response({
            "par_province": list(stats_province),
            "par_genre": list(stats_genre),
            "par_niveau_etude": list(stats_niveau),
            "total": retenues.count(),
        })


class ExportCandidaturesRetenuesExcelView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        # Récupère les stats comme dans CandidaturesRetenuesStatsView
        retenues = Candidature.objects.filter(statut='valide')
        stats_province = (
            retenues
            .values('candidat__profil_candidat__province_de_residence')
            .annotate(total=Count('id'))
            .order_by('candidat__province_de_residence')
        )
        stats_genre = (
            retenues
            .values('candidat__profil_candidat__genre')
            .annotate(total=Count('id'))
            .order_by('candidat__profil_candidat__genre')
        )
        stats_niveau = (
            retenues
            .values('candidat__profil_candidat__niveau_etude')
            .annotate(total=Count('id'))
            .order_by('candidat__profil_candidat__niveau_etude')
        )

        # Création du fichier Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stats Candidatures Retenues Par Province"

        # Province
        ws.append(["Statistiques par province"])
        ws.append(["Province", "Total"])
        for row in stats_province:
            ws.append([row['candidat__profil_candidat__province_de_residence'], row['total']])
        ws.append([])

        # Genre
        ws.append(["Statistiques par genre"])
        ws.append(["Genre", "Total"])
        for row in stats_genre:
            ws.append([row['candidat__profil_candidat__genre'], row['total']])
        ws.append([])

        # Niveau d'étude
        ws.append(["Statistiques par niveau d'étude"])
        ws.append(["Niveau d'étude", "Total"])
        for row in stats_niveau:
            ws.append([row['candidat__profil_candidat__niveau_etude'], row['total']])
        ws.append([])

        # Total général
        ws.append(["Total général", retenues.count()])

        # Préparation de la réponse HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=stats_candidatures_retenues.xlsx'
        wb.save(response)
        return response


class CandidaturesNonRetenuesStatsView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        # Filtre les candidatures non retenues
        non_retenues = Candidature.objects.filter(statut='rejete')

        return Response({
            "liste": non_retenues,
            "total": non_retenues.count()
        })


class ExportCandidaturesNonRetenuesExcelView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        non_retenues = Candidature.objects.filter(statut='rejete')
        stats_province = (
            non_retenues
            .values('candidat__profil_candidat__province_residence')
            .annotate(total=Count('id'))
            .order_by('candidat__profil_candidat__province_residence')
        )
        stats_genre = (
            non_retenues
            .values('candidat__profil_candidat__genre')
            .annotate(total=Count('id'))
            .order_by('candidat__profil_candidat__genre')
        )
        stats_niveau = (
            non_retenues
            .values('candidat__profil_candidat__niveau_etude')
            .annotate(total=Count('id'))
            .order_by('candidat__profil_candidat__niveau_etude')
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stats Candidatures Non Retenues"

        # Province
        ws.append(["Statistiques par province"])
        ws.append(["Province", "Total"])
        for row in stats_province:
            ws.append([row['candidat__profil_candidat__province_residence'], row['total']])
        ws.append([])

        # Genre
        ws.append(["Statistiques par genre"])
        ws.append(["Genre", "Total"])
        for row in stats_genre:
            ws.append([row['candidat__profil_candidat__genre'], row['total']])
        ws.append([])

        # Niveau d'étude
        ws.append(["Statistiques par niveau d'étude"])
        ws.append(["Niveau d'étude", "Total"])
        for row in stats_niveau:
            ws.append([row['candidat__profil_candidat__niveau_etude'], row['total']])
        ws.append([])

        # Total général
        ws.append(["Total général", non_retenues.count()])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=stats_candidatures_non_retenues.xlsx'
        wb.save(response)
        return response


class TraiterCandidatureView(RetrieveUpdateAPIView):
    queryset = Candidature.objects.all()
    permission_classes = [IsAdminUser]
    serializer_class = TraiterCandidatureSerializer
    #authentication_classes = (JWTAuthentication,)

    def get_object(self):
        self.DOCUMENTS = ["cv", "lettre_motivation", "titre_academique", "piece_identite", "aptitude_physique"]

    def evaluer_candidature(self, candidature, user, annotations):
        """
        annotations: dict du type {
            "cv": {"commentaire": "conforme", "note": 10},
            "lettre_motivation": {"commentaire": "non conforme", "note": 7},
            ...
        }
        """
        # 1. Enregistre le traitement
        Traitement.objects.create(
            candidature=candidature,
            evaluateur=user,
            document=annotations,
            decision="",  # à compléter selon ton besoin
            observations=""
        )

        # 2. Récupère tous les traitements existants pour cette candidature
        traitements = Traitement.objects.filter(candidature=candidature).order_by('date_traitement')
        if traitements.count() == 1:
            # Premier évaluateur : passage à "en_traitement"
            candidature.statut = "en_traitement"
            candidature.save(update_fields=["statut"])
            return {"detail": "Traitement du premier évaluateur enregistré."}

        elif traitements.count() == 2:
            # Deuxième évaluateur : calcul de la moyenne
            notes_eval1 = traitements[0].document
            notes_eval2 = traitements[1].document

            discordance = False
            total_eval1 = 0
            total_eval2 = 0
            for doc in self.DOCUMENTS:
                note1 = notes_eval1.get(doc, {}).get("note", 0)
                note2 = notes_eval2.get(doc, {}).get("note", 0)
                total_eval1 += note1
                total_eval2 += note2
                # Discordance si écart >= 3 points sur un document
                if abs(note1 - note2) >= 3:
                    discordance = True

            moyenne = mean([total_eval1, total_eval2])

            if discordance:
                candidature.statut = "en_conflit"
                candidature.save(update_fields=["statut"])
                return {"detail": "Discordance détectée, passage à un troisième évaluateur."}

            # Décision finale
            if moyenne == 50:
                candidature.statut = "valide"
            else:
                candidature.statut = "rejete"
            candidature.save(update_fields=["statut"])
            return {"detail": f"Traitement terminé. Moyenne: {moyenne}. Statut: {candidature.statut}"}

        elif traitements.count() == 3:
            # Troisième évaluateur : départage
            notes = [t.document for t in traitements]
            totaux = []
            for t in notes:
                total = sum(t.get(doc, {}).get("note", 0) for doc in self.DOCUMENTS)
                totaux.append(total)
            moyenne = mean(totaux)
            if moyenne == 50:
                candidature.statut = "valide"
            else:
                candidature.statut = "rejete"
            candidature.save(update_fields=["statut"])
            return {"detail": f"Départage effectué. Moyenne: {moyenne}. Statut: {candidature.statut}"}

        else:
            return {"detail": "Nombre d'évaluations inattendu."}

    # ...existing code...

    # Exemple d'utilisation dans la méthode post de TraiterCandidatureView :
    def post(self, request, pk):
        try:
            candidature = Candidature.objects.get(id=pk)
        except Candidature.DoesNotExist:
            return Response({"detail": "Candidature introuvable."}, status=404)

        user = request.user
        annotations = request.data.get("annotations", {})
        resultat = self.evaluer_candidature(candidature, user, annotations)
        return Response(resultat)
    # ...existing code...

    def patch(self, request, pk):
        try:
            candidature = Candidature.objects.get(pk=pk)
        except Candidature.DoesNotExist:
            return Response({"detail": "Candidature introuvable."}, status=status.HTTP_404_NOT_FOUND)

        data = request.data
        serializer = CandidatureAddSerializer(candidature, data=data, partial=True)
        if serializer.is_valid():
            serializer.save(traite_par=request.user)

            n = Notification.objects.create(
                message="Votre candidature a été traitée. Vous serez fixé au moment de la publication." if candidature.traite else "Votre candidature a été mise à jour.",
                link=f"/candidature/{candidature.id}/"
            )
            user = candidature.candidat
            n.users.set([user])

            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    def post(self, request, pk):
        try:
            candidature = Candidature.objects.get(pk=pk)
        except Candidature.DoesNotExist:
            return Response({"detail": "Candidature introuvable."}, status=404)

        user = request.user
        # Détermine l'étape selon le rôle de l'utilisateur
        if user.role == "evaluateur1":
            #etape = "evaluateur1"
            statut = "en_traitement"
        elif user.role == "evaluateur2":
            #etape = "evaluateur2"
            statut = "traite"
        else:
            return Response({"detail": "Rôle non autorisé."}, status=403)

        # Crée un traitement lié à cette candidature et cet évaluateur
        Traitement.objects.create(
            candidature=candidature,
            evaluateur=user,
            document=request.data.get("document", ""),
            decision=request.data.get("decision", ""),
            observations=request.data.get("observations", "")
        )

        # Met à jour la candidature si besoin (optionnel)
        candidature.statut = statut
        candidature.save(update_fields=["statut"])

        return Response({"detail": "Traitement enregistré."})


class EvaluerCandidatureView(CreateAPIView):
    """
    Vue permettant à un évaluateur d'évaluer une candidature.
    Utilise le serializer TraitementSerializer.
    Ajoute un enregistrement de traitement pour la candidature.
    Vérifie que l'évaluateur n'a pas déjà évalué cette candidature.
    """
    serializer_class = TraitementSerializer
    permission_classes = [IsEvaluator]
    authentication_classes = [JWTAuthentication]

    def perform_create (self, serializer):
        candidature_id = self.request.data.get('candidature_id')
        
        if not candidature_id:
            return Response ({'candidature': 'Ce champ est requis.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            candidature_retournee = Candidature.objects.get(id=candidature_id) if candidature_id else None
        except Candidature.DoesNotExist:
            raise DRFValidationError({'candidature': 'Candidature introuvable.'})

        # Vérifie que l'évaluateur n'a pas déjà évalué cette candidature
        if Traitement.objects.filter(candidature=candidature_retournee, evaluateur=self.request.user).exists():
            return Response ({'candidature': 'Vous avez déjà évalué cette candidature.'}, status=status.HTTP_400_BAD_REQUEST)

        serializer.save(
            candidature=candidature_retournee,
            evaluateur=self.request.user
        )
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)


class HistoriqueTraitementsView(ListAPIView):
    serializer_class = TraitementListSerializer
    permission_classes = [IsAuthenticated]
    queryset = Traitement.objects.all()

    # def get_queryset(self):
        # candidature_id = self.kwargs['pk']
        # return Traitement.objects.filter(candidature_id=candidature_id).order_by('-date_traitement')


class CandidaturesATraiterView(ListAPIView):
    serializer_class = CandidatureAddSerializer
    permission_classes = [IsAdminUser]
    #authentication_classes = (JWTAuthentication,)

    def get_queryset(self):
        return Candidature.objects.filter(statut__in=['envoye', 'en_traitement'])
 
################################### RECOURS ###############################
class RecoursViewSet(viewsets.ModelViewSet):
    """
    Permet :
        (1) au candidat de gérer son recours : création, modification, soumission.
        (2) à l'administrateur de consulter les recours.
    """

    queryset = Recours.objects.all()
    serializer_class = RecoursSerializer
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsCandidat | IsAdminUser]

    def get_queryset(self):
        user = self.request.user
        if user.is_staff or user.is_superuser:
            return Recours.objects.all()
        return Recours.objects.filter(candidature__candidat=user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    def perform_update(self, serializer):
        instance = serializer.instance
        if not instance.traite:
            return Response({'restriction': 'Ce recours ne peut plus être modifiée.'}, status=403)
        serializer.save()

    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()
        if not instance.traite:
            return Response({'restriction': 'Ce recours ne peut plus être modifiée.'}, status=403)
        return super().partial_update(request, *args, **kwargs)


class RecoursStatusView(RetrieveAPIView):
    """
    Vue pour récupérer le statut du recours.
    """
    serializer_class = RecoursStatusSerializer
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get_object(self):
        return Recours.objects.get(user=self.request.user)
    

class AjoutRecoursView(CreateAPIView):
    queryset = Recours.objects.all()
    serializer_class = RecoursSerializer
    permission_classes = [IsCandidat]
    #authentication_classes = (JWTAuthentication,)

    def perform_create(self, serializer):
        # Automatically set the current user as the candidat
        user = self.request.user
        if not user.role == 'candidat':
            # on vérifie si l'utilisateur a le role 'candidat'
            # If not, return a 403 Forbidden response
            return Response({'Erreur': 'Cette opération requiert uniquement les privilèges Candidat.'}, status=403)
        serializer.save(candidat=user)


class RecoursListView(ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = RecoursSerializer
    authentication_classes = (JWTAuthentication,)

    def get_queryset(self):
        user = self.request.user
        if user.is_staff or user.is_superuser:
            return Recours.objects.all()
        return Recours.objects.filter(candidature__candidat=user)

    def perform_create(self, serializer):
        serializer.save()

    @action(detail=False, methods=['get'], permission_classes=[permissions.IsAdminUser])
    def non_traite(self, request):
        queryset = Recours.objects.filter(traite=False)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    

class TraiterRecoursView(APIView):
    permission_classes = [IsAdminUser]
    #authentication_classes = (JWTAuthentication,)
    serializer_class = RecoursSerializer

    def patch(self, request, pk):
        try:
            recours = Recours.objects.get(pk=pk)
        except Recours.DoesNotExist:
            return Response({"detail": "Recours introuvable."}, status=status.HTTP_404_NOT_FOUND)

        data = request.data
        old_statut = recours.traite
        # Met à jour les champs pertinents
        recours.traite = True
        recours.commentaire_admin = data.get('commentaire_admin', recours.commentaire_admin)
        recours.date_traitement = now()
        recours.traite_par = request.user
        recours.save()

        # Notification lors du traitement
        send_mail(
            subject="Votre recours a été traité",
            message=f"Bonjour, votre recours a été traité. Commentaire : {recours.commentaire_admin}",
            from_email="no-reply@ena.com",
            recipient_list=[recours.candidature.candidat.email],
            fail_silently=True,
        )

        n = Notification.objects.create(
            message="Votre recours a été traité. Statut : validé." if recours.traite else "Votre recours a été mis à jour.",
            link=f"/recours/{recours.id}/"
        )
        user=recours.candidature.candidat
        n.users.set([user])

        
        # Historique
        RecoursActionHistory.objects.create(
            recours=recours,
            admin=request.user,
            action="Traité" if not old_statut else "Mis à jour",
            commentaire=recours.commentaire_admin
        )

        serializer = RecoursSerializer(recours)
        return Response(serializer.data, status=status.HTTP_200_OK)
    

class RecoursHistoryView(ListAPIView):
    serializer_class = RecoursActionHistorySerializer
    permission_classes = [IsAdminUser]
    #authentication_classes = (JWTAuthentication,)

    def get_queryset(self):
        recours_id = self.kwargs['pk']
        return RecoursActionHistory.objects.filter(recours_id=recours_id).order_by('-date_action')

################################# TRAINING MODULES ###############################
class TrainingModuleListView(ListAPIView):
    queryset = TrainingModule.objects.all()
    serializer_class = TrainingModuleSerializer
    permission_classes = [IsAuthenticated]


class TrainingModuleDetailView(RetrieveAPIView):
    queryset = TrainingModule.objects.all()
    serializer_class = TrainingModuleSerializer
    permission_classes = [IsAuthenticated]


class SubmitQuizAnswerView(CreateAPIView):
    serializer_class = QuizAnswerSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class UserQuizAnswersView(ListAPIView):
    serializer_class = QuizAnswerSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return QuizAnswer.objects.filter(user=self.request.user)


class UserModuleProgressView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, module_id):
        module = TrainingModule.objects.get(pk=module_id)
        total_questions = module.questions.count()
        answered = QuizAnswer.objects.filter(user=request.user, question__module=module).count()
        correct = QuizAnswer.objects.filter(user=request.user, question__module=module, is_correct=True).count()
        return Response({
            "module": module.title,
            "total_questions": total_questions,
            "answered": answered,
            "correct": correct,
            "progress_percent": (answered / total_questions * 100) if total_questions else 0,
            "score_percent": (correct / total_questions * 100) if total_questions else 0,
        })
    

class TrainingModulesProgressView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        modules = TrainingModule.objects.all()
        progress_list = []
        for module in modules:
            total_questions = module.questions.count()
            answered = QuizAnswer.objects.filter(user=request.user, question__module=module).count()
            correct = QuizAnswer.objects.filter(user=request.user, question__module=module, is_correct=True).count()
            progress_list.append({
                "module_id": str(module.id),
                "module_title": module.title,
                "total_questions": total_questions,
                "answered": answered,
                "correct": correct,
                "progress_percent": (answered / total_questions * 100) if total_questions else 0,
                "score_percent": (correct / total_questions * 100) if total_questions else 0,
            })
        return Response(progress_list)
    
#List Candidate's Quiz Answers
class CandidateQuizAnswersView(ListAPIView):
    serializer_class = QuizAnswerSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return QuizAnswer.objects.filter(user=self.request.user)
    

#List Modules started by the candidate
class CandidateStartedModulesView(ListAPIView):
    serializer_class = TrainingModuleSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        # Modules where the candidate has answered at least one question
        module_ids = QuizAnswer.objects.filter(user=self.request.user).values_list('question__module', flat=True).distinct()
        return TrainingModule.objects.filter(id__in=module_ids)


def get_question_options(request, question_id):
    try:
        q = QuizQuestion.objects.get(pk=question_id)
        return JsonResponse({
            'A': q.option_a,
            'B': q.option_b,
            'C': q.option_c,
            'D': q.option_d,
        })
    except QuizQuestion.DoesNotExist:
        return JsonResponse({}, status=404)
    
